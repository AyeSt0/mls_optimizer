
#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
12_llm_translate.py  (per-row autosave + persistent cache + WAL)

- Reads Excel: col1 RU, col2 speaker, col3 EN, col4 CN(existing), writes col5 OUT
- Only translate where col5 is empty (resume-friendly)
- Per-row WAL append to --wal-file (JSONL), ALWAYS
- Persistent cache (--cache-file JSONL) keyed by hash(ru|en|speaker|ctx|glossary_digest|target)
- Longest-first glossary mapping
- Optional immediate Excel write per row (autosave-every=1 or autosave-seconds=0)
- Minimal dynamic throttling by rpm/tpm (best-effort)

NOTE: This is a focused drop-in. It accepts a superset of args used in your GUI.
"""

import argparse
import sys
import json
import os
import time
import math
import re
import queue
import threading
from pathlib import Path
from datetime import datetime
import hashlib

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- Settings helpers ------------------------------------------------------

def _read_yaml_dict(p: Path):
    try:
        import yaml
    except Exception:
        return {}
    if not p.exists():
        return {}
    with p.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def _settings():
    base = Path("config")
    local = base / "settings.local.yaml"
    default = base / "settings.yaml"
    s = {}
    s.update(_read_yaml_dict(default))
    s.update(_read_yaml_dict(local))
    return s

def _ensure_v1(url: str) -> str:
    if url.endswith("/v1"):
        return url
    if url.endswith("/"):
        return url + "v1"
    return url + "/v1"

# --- Glossary --------------------------------------------------------------

def load_glossary(glossary_path: str):
    if not glossary_path:
        return [], "no-glossary"
    p = Path(glossary_path)
    if not p.exists():
        return [], "no-glossary"
    data = json.loads(p.read_text("utf-8"))
    # data may be: {"Professor Richardson": "理查森教授", ...}
    items = []
    for k, v in data.items():
        if not k or not v:
            continue
        items.append((k, v))
    # longest-first by source length (desc)
    items.sort(key=lambda kv: len(kv[0]), reverse=True)
    digest = hashlib.sha1(json.dumps(data, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()
    return items, digest

def apply_glossary(s: str, gls_items):
    if not s or not gls_items:
        return s
    # longest-first exact phrase replacement; respect word boundaries for latin tokens
    out = s
    for src, tgt in gls_items:
        try:
            # If purely ASCII token, do boundary-aware replace, else plain replace
            if all(ord(ch) < 128 for ch in src):
                pat = r'(?<![\w])' + re.escape(src) + r'(?![\w])'
                out = re.sub(pat, tgt, out)
            else:
                out = out.replace(src, tgt)
        except re.error:
            out = out.replace(src, tgt)
    return out

# --- Cache (JSONL) ---------------------------------------------------------

class JsonlCache:
    def __init__(self, path: Path):
        self.path = path
        self.lock = threading.Lock()
        self.map = {}  # key -> value
        if path.exists():
            for line in path.open("r", encoding="utf-8"):
                if not line.strip():
                    continue
                try:
                    row = json.loads(line)
                    self.map[row["key"]] = row["val"]
                except Exception:
                    continue

    def get(self, key):
        return self.map.get(key)

    def put(self, key, val):
        with self.lock:
            self.map[key] = val
            with self.path.open("a", encoding="utf-8") as f:
                f.write(json.dumps({"key": key, "val": val}, ensure_ascii=False) + "\n")

# --- WAL -------------------------------------------------------------------

class Wal:
    def __init__(self, path: Path):
        self.path = path
        self.lock = threading.Lock()

    def append(self, rec: dict):
        rec2 = dict(rec)
        rec2["ts"] = datetime.utcnow().isoformat()
        line = json.dumps(rec2, ensure_ascii=False)
        with self.lock:
            with self.path.open("a", encoding="utf-8") as f:
                f.write(line + "\n")

# --- Excel writer (single-thread) -----------------------------------------

class ExcelWriterThread(threading.Thread):
    def __init__(self, excel_path: Path, sheet_name=None, flush_every=1, flush_seconds=1.0):
        super().__init__(daemon=True)
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.flush_every = max(1, int(flush_every))
        self.flush_seconds = max(0.0, float(flush_seconds))
        self.q = queue.Queue()
        self._stop = threading.Event()
        self._last_save = time.time()
        self._pending = 0

        self.wb = load_workbook(str(excel_path))
        if sheet_name and sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]

    def run(self):
        while not self._stop.is_set():
            try:
                item = self.q.get(timeout=0.2)
            except queue.Empty:
                item = None

            if item:
                row_idx, col_idx, val = item
                try:
                    self.ws.cell(row=row_idx, column=col_idx, value=val)
                except Exception:
                    pass
                self._pending += 1

            now = time.time()
            if self._pending >= self.flush_every or (self.flush_seconds and (now - self._last_save) >= self.flush_seconds):
                try:
                    self.wb.save(str(self.excel_path))
                    self._last_save = now
                    self._pending = 0
                except Exception:
                    # ignore save errors, try next round
                    pass

        # Final save on stop
        try:
            self.wb.save(str(self.excel_path))
        except Exception:
            pass
        try:
            self.wb.close()
        except Exception:
            pass

    def submit(self, row_idx, col_idx, val):
        self.q.put((row_idx, col_idx, val))

    def stop(self):
        self._stop.set()

# --- LLM client (OpenAI compatible) ----------------------------------------

def build_client(provider: str, cfg: dict):
    from openai import OpenAI

    prov = (provider or "").lower()
    llm_cfg = cfg.get("llm") or cfg.get("model") or {}
    if prov in ("deepseek", "siliconflow"):
        c = (llm_cfg.get("deepseek") or {})
        api_key = os.getenv("DEEPSEEK_API_KEY") or os.getenv("SILICONFLOW_API_KEY") or c.get("api_key")
        base = _ensure_v1((c.get("base_url") or "https://api.siliconflow.cn"))
        model = c.get("name") or c.get("model") or "deepseek-ai/DeepSeek-V3.2-Exp"
        if not api_key:
            raise RuntimeError("Missing DeepSeek/SiliconFlow API key.")
        client = OpenAI(api_key=api_key, base_url=base)
        return client, model
    # default openai
    oc = (llm_cfg.get("openai") or {})
    api_key = os.getenv("OPENAI_API_KEY") or oc.get("api_key")
    base = _ensure_v1((oc.get("base_url") or "https://api.openai.com"))
    model = oc.get("name") or oc.get("model") or "gpt-4o-mini"
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY.")
    client = OpenAI(api_key=api_key, base_url=base)
    return client, model

# --- Prompt ----------------------------------------------------------------

def make_prompt(target_lang_label: str, ru: str, en: str, speaker: str, scene_hint: str, glossary_items):
    # High-signal instruction tuned for Sunshine Town college VN
    sys_inst = (
        "You are a senior localization writer for an adult college visual novel set in Sunshine Town. "
        "All characters are 18+. Keep seductive, natural, immersive tone for Chinese players. "
        "Use EN for literal meaning; use RU for mood/sensual nuance. "
        "Keep placeholders exactly as-is: [mcname], [mcsurname], mcname, [var], {{tag}}, <tag>.\n"
        "If speaker seems like a UI label or location list (often 'string'), prefer concise noun phrases; "
        "for school locations, translate to 'X教室/办公室/更衣室/体育馆/医务室/礼堂/学院大门' etc.\n"
        "For brand names (Patreon, Instagram, Lovense) keep original. Technical: Sync ID, Local IP, Port -> keep glossary mapping.\n"
        "Never output quotes or explanations. Output ONLY the final Chinese. "
        "Keep explicit erotic content as-is (no censorship). "
    )
    if glossary_items:
        # add top-N glossary lines to bias the model
        tips = "术语表（优先长词匹配）：\n" + "\n".join([f"- {k} -> {v}" for k,v in glossary_items[:50]])
    else:
        tips = ""

    # style hints by speaker
    style = ""
    sp = (speaker or "").lower()
    if sp in ("principal_richardson","principal_richardson_t","teacher","teacher_adams","teacher_clark","teacher_hill","teacher_morris","librarian_wilson"):
        style = "语气：学术/权威，但自然口语。"
    elif sp in ("trainer_brooks","coach","coach_brooks","trainer_brooks_t"):
        style = "语气：运动指导口吻，直接有力。"
    elif sp in ("mrs","lady","madam","madame","secretary_young","emilys_mother","emilys_mother_t"):
        style = "语气：礼貌克制，但自然口语（不必强制使用‘您’）。"
    elif sp == "string":
        style = "语气：UI/地点短语，名词化，避免清单式断句。"
    else:
        style = "语气：自然、口语化、带情绪张力；对话不要清单式直译。"

    user = f"""\
[目标语言] {target_lang_label}

[角色] speaker = {speaker}
[场景提示] {scene_hint or "（连续剧情场景）"}

[英文原文]
{en or ""}

[俄文原文]
{ru or ""}

[{style}]
{tips}
"""
    return sys_inst, user

# --- Rate / token helpers (best-effort) ------------------------------------

class SimpleRate:
    def __init__(self, rpm=60, tpm_max=100000):
        self.rpm = max(1, int(rpm))
        self.tpm_max = max(1000, int(tpm_max))
        self.lock = threading.Lock()
        self.req_times = []  # timestamps of last 60s
        self.tokens_in_60 = 0

    def admit(self, est_tokens=400):
        # crude pacing: ensure <= rpm in last 60s and token budget roughly respected
        while True:
            with self.lock:
                now = time.time()
                self.req_times = [t for t in self.req_times if now - t < 60.0]
                if len(self.req_times) < self.rpm and self.tokens_in_60 + est_tokens < self.tpm_max:
                    self.req_times.append(now)
                    self.tokens_in_60 += est_tokens
                    break
            time.sleep(0.05)

    def decay(self):
        with self.lock:
            now = time.time()
            old_len = len(self.req_times)
            self.req_times = [t for t in self.req_times if now - t < 60.0]
            if len(self.req_times) != old_len:
                # rough token decay
                self.tokens_in_60 = int(self.tokens_in_60 * (len(self.req_times) / max(1, old_len)))

# --- Main ------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Path to XLSX")
    ap.add_argument("--sheet", default=None, help="Sheet name (optional)")
    ap.add_argument("--target-lang", default="zh-CN")
    ap.add_argument("--provider", default=None, help="deepseek/openai; default from settings")
    ap.add_argument("--settings", default="config/settings.local.yaml")
    ap.add_argument("--glossary", default=None)
    ap.add_argument("--rpm", type=int, default=60)
    ap.add_argument("--tpm-max", type=int, default=100000)
    ap.add_argument("--autosave-every", type=int, default=1, help="Rows per Excel save (1 = per-row)")
    ap.add_argument("--autosave-seconds", type=float, default=0.0, help="Seconds between Excel saves (0 = every row)")
    ap.add_argument("--checkpoint-file", default="artifacts/ckpt.translate.jsonl")
    ap.add_argument("--wal-file", default="artifacts/translate.wal.jsonl")
    ap.add_argument("--cache-file", default="artifacts/cache.translate.jsonl")
    ap.add_argument("--out", default=None)
    ap.add_argument("--show-lines", action="store_true")
    ap.add_argument("--max-chars", type=int, default=120)
    args = ap.parse_args()

    cfg = _settings()
    provider = (args.provider or cfg.get("provider") or "deepseek")

    client, model = build_client(provider, cfg)
    print(f"[CFG] provider={provider} model={model} rpm={args.rpm} tpm={args.tpm_max}", flush=True)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise SystemExit(f"Excel not found: {excel_path}")

    df = pd.read_excel(str(excel_path), sheet_name=args.sheet)
    # Expect columns 1..5: ru, speaker, en, cn, out
    # use positional access for robustness
    def _get(row, idx):
        try:
            return str(row[idx]) if not (isinstance(row[idx], float) and math.isnan(row[idx])) else ""
        except Exception:
            try:
                return str(row.iloc[idx])
            except Exception:
                return ""

    # Prepare output column if not exists
    if df.shape[1] < 5:
        # add 5th column
        while df.shape[1] < 5:
            df[df.shape[1]] = ""
    # rows to translate: col5 empty
    mask_empty = df.iloc[:,4].isna() | (df.iloc[:,4].astype(str).str.strip() == "")
    to_translate_idx = list(df[mask_empty].index.values.tolist())

    print(f"[PLAN] total={len(df)} empty(col5)={mask_empty.sum()} to_translate={len(to_translate_idx)}", flush=True)

    gls_items, gls_digest = load_glossary(args.glossary) if args.glossary else ([], "no-glossary")
    cache = JsonlCache(Path(args.cache_file))
    wal = Wal(Path(args.wal_file))
    rate = SimpleRate(args.rpm, args.tpm_max)

    # Excel writer thread
    writer = ExcelWriterThread(excel_path, sheet_name=args.sheet,
                               flush_every=max(1, args.autosave_every),
                               flush_seconds=max(0.0, args.autosave_seconds))
    writer.start()

    # Optional: a simple scene hint using previous/next lines; keep cheap
    def scene_hint_for(i):
        ru_prev = df.iloc[i-1,0] if i>0 else ""
        en_prev = df.iloc[i-1,2] if i>0 else ""
        ru_next = df.iloc[i+1,0] if i+1<len(df) else ""
        en_next = df.iloc[i+1,2] if i+1<len(df) else ""
        parts = []
        if isinstance(en_prev, str) and en_prev.strip():
            parts.append(f"Prev: {en_prev}")
        if isinstance(en_next, str) and en_next.strip():
            parts.append(f"Next: {en_next}")
        return " | ".join(parts)[:200]

    # Main loop (sequential for reliability; per-row autosave + cache give robustness)
    success = 0
    last_save = time.time()
    for i in to_translate_idx:
        ru = str(df.iloc[i,0]) if not pd.isna(df.iloc[i,0]) else ""
        sp = str(df.iloc[i,1]) if not pd.isna(df.iloc[i,1]) else ""
        en = str(df.iloc[i,2]) if not pd.isna(df.iloc[i,2]) else ""

        ctx = scene_hint_for(i)
        key_src = json.dumps({"ru":ru, "en":en, "sp":sp, "ctx":ctx, "gls":gls_digest, "t":args.target_lang}, ensure_ascii=False)
        key = hashlib.sha1(key_src.encode("utf-8")).hexdigest()

        cached = cache.get(key)
        if cached:
            out = cached
            df.iat[i,4] = out
            # per-row WAL & Excel
            wal.append({"row": int(i), "speaker": sp, "ru": ru, "en": en, "out": out, "cached": True})
            # Excel row index: +2 because pandas assumes header row; our file likely has header row
            writer.submit(i+2, 5, out)
            success += 1
            if args.show_lines:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] (cache) row={i} speaker={sp}")
            continue

        sys_inst, user_msg = make_prompt(args.target_lang, ru, en, sp, ctx, gls_items)

        # pacing
        est_tokens = max(200, min(800, len(ru)//2 + len(en)))  # crude estimate
        rate.admit(est_tokens)

        try:
            rsp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role":"system","content":sys_inst},
                    {"role":"user","content":user_msg},
                ],
                temperature=0.2,
                top_p=0.9,
                max_tokens=300,
            )
            out = (rsp.choices[0].message.content or "").strip()
        except Exception as e:
            print(f"[WARN] API error row={i}: {e}", flush=True)
            time.sleep(1.0)
            continue

        # Apply glossary post-fix too (belt and suspenders)
        out2 = apply_glossary(out, gls_items)
        df.iat[i,4] = out2
        cache.put(key, out2)

        # per-row WAL & Excel
        wal.append({"row": int(i), "speaker": sp, "ru": ru, "en": en, "out": out2, "cached": False})
        writer.submit(i+2, 5, out2)
        success += 1

        if args.show_lines:
            try:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] RU: {ru[:80]}")
                print(f"[{datetime.now().strftime('%H:%M:%S')}] EN: {en[:80]}")
            except Exception:
                pass
            print(f"[{datetime.now().strftime('%H:%M:%S')}] OUT: {out2[:80]}")
            print(f"[SEND] row={i} speaker={sp}")
        print(f"[PROGRESS] {success}/{len(to_translate_idx)}", flush=True)

        rate.decay()

    # Final save snapshot
    writer.stop()
    # Also write a final synthesized output file if requested
    out_path = args.out or (excel_path.parent / f"{excel_path.stem}.{args.target_lang}.llm.xlsx")
    try:
        # reload original workbook to ensure all pending writes are persisted
        # but the writer.save() already did; here we just re-save under a new name
        from shutil import copyfile
        copyfile(excel_path, out_path)
        print(f"[OK] Done. Output -> {out_path}", flush=True)
    except Exception:
        print(f"[OK] Done. (kept in-place) -> {excel_path}", flush=True)

if __name__ == "__main__":
    main()
